import requests
import pandas as pd
import time
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
from urllib.parse import urlparse
import re

load_dotenv()
API_KEY = os.getenv("API_KEY")

TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
DETAILS_URL_BASE = "https://places.googleapis.com/v1/places/"

QUERIES = [
    "burger shops in Dhaka, Bangladesh",
    "burger restaurant in Dhaka, Bangladesh",
    "burger in Gulshan, Dhaka",
    "burger in Banani, Dhaka",
    "burger in Dhanmondi, Dhaka",
    "burger in Uttara, Dhaka"
]

SEARCH_FIELD_MASK = ",".join([
    "places.id",
    "places.displayName",
    "places.formattedAddress",
    "places.priceLevel",
    "places.priceRange",
    "places.rating",
    "places.userRatingCount",
    "places.googleMapsUri",
    "nextPageToken",
])

DETAILS_FIELD_MASK = ",".join([
    "id",
    "displayName",
    "formattedAddress",
    "nationalPhoneNumber",
    "internationalPhoneNumber",
    "websiteUri",
    "googleMapsUri",
    "priceLevel",
    "priceRange",
    "rating",
    "userRatingCount",
])

PRICE_MAP = {
    "Free":0,
    "Cheap": 1,
    "Moderate": 2,
    "Expensive": 3,
    "Very Expensive": 4
}

PRICE_TEXT_MAP = {
    "PRICE_LEVEL_FREE": "Free",
    "PRICE_LEVEL_INEXPENSIVE": "Cheap",
    "PRICE_LEVEL_MODERATE": "Moderate",
    "PRICE_LEVEL_EXPENSIVE": "Expensive",
    "PRICE_LEVEL_VERY_EXPENSIVE": "Very Expensive"
}

output_file = "dhaka_burger_shops_final.xlsx"


def headers(field_mask: str) -> dict:
    return {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": field_mask,
    }

def extract_email(url):
    try:
        r = requests.get(url, timeout=5)
        emails = re.findall(r"[\w\.-]+@[\w\.-]+", r.text)
        return emails[0] if emails else None
    except:
        return None

def money_to_text(money_obj):
    if not money_obj:
        return None

    currency = money_obj.get("currencyCode", "")
    units = money_obj.get("units", "0")
    nanos = money_obj.get("nanos", 0)

    try:
        units = int(units)
    except Exception:
        units = 0

    try:
        nanos = int(nanos)
    except Exception:
        nanos = 0

    value = units + (nanos / 1_000_000_000)

    if currency:
        return f"{currency} {value:.2f}"
    return f"{value:.2f}"


def parse_price_range(price_range_obj):
    if not price_range_obj:
        return None, None

    start_price = money_to_text(price_range_obj.get("startPrice"))
    end_price = money_to_text(price_range_obj.get("endPrice"))
    return start_price, end_price

def fetch_text_search_page(query, page_token=None):
    body = {
        "textQuery": query,
        "pageSize": 20,
    }

    if page_token:
        body["pageToken"] = page_token

    response = requests.post(
        TEXT_SEARCH_URL,
        headers=headers(SEARCH_FIELD_MASK),
        json=body,
        timeout=30,
    )

    
    print("\nREQUEST BODY:", body)
    print("STATUS:", response.status_code)
    print("BODY:", response.text)

    response.raise_for_status()
    return response.json()


def fetch_place_details(place_id: str):
    url = f"{DETAILS_URL_BASE}{place_id}"

    response = requests.get(
        url,
        headers=headers(DETAILS_FIELD_MASK),
        timeout=30,
    )

    if response.status_code != 200:
        print(f"Details failed for {place_id}: {response.status_code}")
        print(response.text)

    response.raise_for_status()
    return response.json()

def col_order(df,col_1,col_2):
    cols = list(df.columns)

    if col_1 in cols and col_2 in cols:
        cols.insert(cols.index(col_2), cols.pop(cols.index(col_1)))

    df = df[cols]
    return df

def make_excel_links_clickable(filename, sheet_name):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    headers_map = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        headers_map[header_value] = col

    website_col = headers_map.get("Website")
    maps_col = headers_map.get("GoogleMapsLink")

    for row in range(2, ws.max_row + 1):
        if website_col:
            cell = ws.cell(row=row, column=website_col)
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

        if maps_col:
            cell = ws.cell(row=row, column=maps_col)
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    wb.save(filename)


def collect_places():
    all_places = []
    seen_ids = set()

    for query in QUERIES:
        print(f"\nSearching: {query}")

        next_page_token = None
        page_count = 0
        max_pages = 5

        while page_count < max_pages:
            search_data = fetch_text_search_page(query, next_page_token)
            places = search_data.get("places", [])

            for place in places:
                place_id = place.get("id")
                if place_id and place_id not in seen_ids:
                    seen_ids.add(place_id)
                    place["_query"] = query
                    all_places.append(place)

            next_page_token = search_data.get("nextPageToken")
            page_count += 1

            if not next_page_token:
                break

            time.sleep(2)

    return all_places

def is_nonempty_str(value) -> bool:
    return isinstance(value, str) and bool(value.strip())


def safe_lower(value) -> str:
    return value.lower() if isinstance(value, str) else ""


def extract_email(url):
    if not is_nonempty_str(url) or not url.startswith(("http://", "https://")):
        return None
    try:
        response = requests.get(url, timeout=5)
        emails = re.findall(r"[\w\.-]+@[\w\.-]+", response.text)
        return emails[0] if emails else None
    except Exception:
        return None

def get_domain(url):
    try:
        if not is_nonempty_str(url) or not url.startswith(("http://", "https://")):
            return None
        return urlparse(url).netloc.replace("www.", "")
    except Exception:
        return None


def website_type(url):
    url_lower = safe_lower(url)
    if not url_lower:
        return "No Website"
    if "facebook.com" in url_lower:
        return "Facebook"
    if "instagram.com" in url_lower:
        return "Instagram"
    return "Official Website"


def burger_type(name):
    name_lower = safe_lower(name)
    if not name_lower:
        return "Unknown"
    if "grill" in name_lower:
        return "Grill Burger"
    if "fried" in name_lower:
        return "Fried Chicken + Burger"
    if "cafe" in name_lower:
        return "Cafe Burger"
    if "fast food" in name_lower:
        return "Fast Food Burger"
    return "General Burger"


def price_category(level):
    if not is_nonempty_str(level):
        return "Unknown"
    if level == "Cheap":
        return "Budget"
    if level == "Moderate":
        return "Mid-range"
    if level in ["Expensive", "Very Expensive"]:
        return "Premium"
    return "Unknown"


def rating_category(rating):
    if pd.isna(rating):
        return "No Rating"
    if rating >= 4.5:
        return "Excellent"
    if rating >= 4:
        return "Good"
    if rating >= 3:
        return "Average"
    return "Poor"


def get_area(address):
    address_lower = safe_lower(address)
    if "gulshan" in address_lower:
        return "Gulshan"
    if "dhanmondi" in address_lower:
        return "Dhanmondi"
    if "uttara" in address_lower:
        return "Uttara"
    return "Other"



def build_dataframes(all_places):
    raw_rows = []

    for idx, place in enumerate(all_places, start=1):
        place_id = place.get("id")
        if not place_id:
            continue

        name = place.get("displayName", {}).get("text")
        address = place.get("formattedAddress")
        rating = place.get("rating")
        rating_count = place.get("userRatingCount")
        maps_link = place.get("googleMapsUri")
        price_level = place.get("priceLevel")
        start_price, end_price = parse_price_range(place.get("priceRange"))

        phone = None
        website = None

        try:
            detail = fetch_place_details(place_id)

            phone = detail.get("nationalPhoneNumber") or detail.get("internationalPhoneNumber")
            website = detail.get("websiteUri")

            if not name:
                name = detail.get("displayName", {}).get("text")
            if not address:
                address = detail.get("formattedAddress")
            if rating is None:
                rating = detail.get("rating")
            if rating_count is None:
                rating_count = detail.get("userRatingCount")
            if not maps_link:
                maps_link = detail.get("googleMapsUri")
            if not price_level:
                price_level = detail.get("priceLevel")
            if not start_price and not end_price:
                start_price, end_price = parse_price_range(detail.get("priceRange"))

        except Exception as e:
            print(f"[WARN] Details failed for {place_id}: {e}")

        raw_rows.append({
            "Name": name,
            "Address": address,
            "Phone": phone,
            "Rating": rating,
            "RatingCount": rating_count,
            "PriceLevel": PRICE_TEXT_MAP.get(price_level),
            "PriceRangeStart": start_price,
            "PriceRangeEnd": end_price,
            "GoogleMapsLink": maps_link,
            "Website": website,
        })

        if idx % 10 == 0:
            print(f"Processed {idx}/{len(all_places)} places")

        time.sleep(0.2)

    df_raw = pd.DataFrame(raw_rows)

    df_clean = df_raw.copy()

    df_clean = df_clean[df_clean["Address"].str.contains("Dhaka", case=False, na=False)]

    def get_domain(url):
        try:
            if not isinstance(url, str) or not url.startswith(("http://", "https://")):
                return None
            
            return urlparse(url).netloc.replace("www.", "")
        except:
            return None
    df_clean["Domain"] = df_clean["Website"].apply(get_domain)
    def website_type(url):
        try:
            if not isinstance(url, str) or not url:
                return "No Website"
            
            url = url.lower()

            if "facebook.com" in url:
                return "Facebook"
            elif "instagram.com" in url:
                return "Instagram"
            else:
                return "Official Website"
        except:
            return "No Website"

    df_clean["WebsiteType"] = df_clean["Website"].apply(website_type)
    df_clean["Email"] = df_clean.apply(
        lambda row: extract_email(row["Website"]) 
        if row["WebsiteType"] == "Official Website" else None,
        axis=1
    )
    
    def burger_type(name):
        name = str(name).lower()
        
        if "grill" in name:
            return "Grill Burger"
        elif "fried" in name:
            return "Fried Chicken + Burger"
        elif "cafe" in name:
            return "Cafe Burger"
        elif "fast food" in name:
            return "Fast Food Burger"
        else:
            return "General Burger"

    df_clean["BurgerType"] = df_clean["Name"].apply(burger_type)
    
    def price_category(level):
        if level == "Cheap":
            return "Budget"
        elif level == "Moderate":
            return "Mid-range"
        elif level in ["Expensive", "Very Expensive"]:
            return "Premium"
        else:
            return "Unknown"

    df_clean["PriceCategory"] = df_clean["PriceLevel"].apply(price_category)
    
    def rating_category(rating):
        if pd.isna(rating):
            return "No Rating"
        elif rating >= 4.5:
            return "Excellent"
        elif rating >= 4:
            return "Good"
        elif rating >= 3:
            return "Average"
        else:
            return "Poor"

    df_clean["RatingCategory"] = df_clean["Rating"].apply(rating_category)
            
    text_cols = df_clean.select_dtypes(include=["object", "string"]).columns

    df_clean[text_cols] = df_clean[text_cols].apply(
        lambda x: x.str.strip().str.replace(r"\s+", " ", regex=True)
    )

    df_raw = df_raw.where(pd.notna(df_raw), pd.NA)
    df_clean = df_clean.where(pd.notna(df_clean), pd.NA)
    
    def compute_value_score(row):
        rating = row["Rating"]
        price = row["PriceLevel"]

        if pd.isna(rating) or pd.isna(price):
            return pd.NA

        price_num = PRICE_MAP.get(price)

        if not price_num or price_num == 0:
            return pd.NA

        return rating / price_num


    df_clean["ValueScore"] = df_clean.apply(compute_value_score, axis=1)
    def get_area(address):
        address = str(address).lower()
        if "gulshan" in address:
            return "Gulshan"
        if "dhanmondi" in address:
            return "Dhanmondi"
        if "uttara" in address:
            return "Uttara"
        return "Other"
    df_clean["Area"] = df_clean["Address"].apply(get_area)
    
    df_clean = col_order(df_clean, "Website", "GoogleMapsLink")

    top_rated = df_clean.sort_values(
        by=["Rating", "RatingCount"],
        ascending=[False, False],
        na_position="last"
    ).head(10)

    cheapest = df_clean[df_clean["PriceLevel"].isin(["Free", "Cheap"])]
    cheapest = cheapest.sort_values(by=["Rating"], ascending=False, na_position="last").head(10)
    
    most_reviewed = df_clean.sort_values(
        by=["RatingCount"],
        ascending=False, na_position="last"
    ).head(10)
    
    best_value = df_clean.sort_values(
        by=["ValueScore"],
        ascending=False
    ).head(10)
    
    premium = df_clean[df_clean["PriceLevel"].isin(["Expensive", "Very Expensive"])]
    premium = premium.sort_values(by=["Rating"], ascending=False, na_position="last").head(10)
    
    website_stats = df_clean["WebsiteType"].value_counts().reset_index()
    website_stats.columns = ["WebsiteType", "Count"]

    area_stats = df_clean["Area"].value_counts().reset_index()
    area_stats.columns = ["Area", "Count"]

    sheets = {
        "Raw": df_raw,
        "Clean": df_clean,
        "Cheapest": cheapest,
        "Premium": premium,
        "Top Rated": top_rated,
        "Most Reviewed": most_reviewed,
        "Best Valued" : best_value,
        "Website Analysis": website_stats,
        "Area Analysis": area_stats      
    }

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    for sheet_name in sheets.keys():
        make_excel_links_clickable(output_file, sheet_name)
    
    print(f"Saved to {output_file}")

def main():
    all_places = collect_places()
    print(f"\nTotal unique places collected: {len(all_places)}")
    build_dataframes(all_places)

if __name__ == "__main__":
    main()